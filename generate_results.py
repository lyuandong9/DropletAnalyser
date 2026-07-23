"""Generate annotated images, Excel statistics, and size distribution plots."""

import json, os, math, numpy as np
from pathlib import Path
from collections import defaultdict
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from tqdm import tqdm

FR = 'H0/50cst-H0110rpm0.00175kw'
DET_DIR = Path('output/detections_v5') / FR
OUT_DIR = Path('output/results') / FR
OUT_DIR.mkdir(parents=True, exist_ok=True)

# === Scale calibration ===
UM_PER_PX = 110.0  # μm per pixel — change this for different experiments

jfiles = sorted([f for f in os.listdir(str(DET_DIR)) if f.endswith('.json')])

# ===== Collect all droplet data =====
all_droplets = []
all_d_eq_px = []

for jf in tqdm(jfiles, desc='Collecting data'):
    img_name = jf.replace('.json', '.bmp')
    with open(DET_DIR / jf) as f:
        dets = json.load(f)
    for d in dets['droplets']:
        area_px = d['area_px']
        d_eq_px = 2.0 * math.sqrt(area_px / math.pi)
        maj, min_ax = d['axes']
        all_d_eq_px.append(d_eq_px)
        all_droplets.append({
            'image': img_name, 'droplet_id': d['id'],
            'center_x_px': d['center'][0], 'center_y_px': d['center'][1],
            'center_x_um': round(d['center'][0] * UM_PER_PX, 1),
            'center_y_um': round(d['center'][1] * UM_PER_PX, 1),
            'area_px': area_px,
            'area_um2': round(area_px * UM_PER_PX * UM_PER_PX, 1),
            'eq_diameter_px': round(d_eq_px, 2),
            'eq_diameter_um': round(d_eq_px * UM_PER_PX, 1),
            'major_axis_px': round(maj, 1),
            'major_axis_um': round(maj * UM_PER_PX, 1),
            'minor_axis_px': round(min_ax, 1),
            'minor_axis_um': round(min_ax * UM_PER_PX, 1),
            'axis_ratio': round(maj / max(min_ax, 0.1), 2),
            'angle_deg': d['angle'],
        })

d_eq_px_arr = np.array(all_d_eq_px)
d_eq_um_arr = d_eq_px_arr * UM_PER_PX
print(f'Total droplets: {len(all_d_eq_px)}')
print(f'D_eq (px): min={d_eq_px_arr.min():.1f} max={d_eq_px_arr.max():.1f} mean={d_eq_px_arr.mean():.1f}')
print(f'D_eq (μm): min={d_eq_um_arr.min():.0f} max={d_eq_um_arr.max():.0f} mean={d_eq_um_arr.mean():.0f} (at {UM_PER_PX} μm/px)')

# ===== Excel =====
wb = Workbook()
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', fgColor='4472C4')
border = Border(bottom=Side(style='thin', color='CCCCCC'))
title_font = Font(bold=True, size=14, color='1F4E79')

ws_summary = wb.active
ws_summary.title = 'Summary'
ws_summary.column_dimensions['A'].width = 30
ws_summary.column_dimensions['B'].width = 22
ws_summary.column_dimensions['C'].width = 22

ws_summary['A1'] = 'Particle Size Analysis Summary'
ws_summary['A1'].font = title_font

# Scale calibration (prominent)
ws_summary['A2'] = 'Scale (μm/px)'
ws_summary['B2'] = UM_PER_PX
ws_summary['B2'].font = Font(bold=True, size=13, color='C00000')
ws_summary['A3'] = 'Folder'; ws_summary['B3'] = FR
ws_summary['A4'] = 'Total Images'; ws_summary['B4'] = len(jfiles)
ws_summary['A5'] = 'Total Droplets'; ws_summary['B5'] = len(all_d_eq_px)

# Summary statistics in both px and μm
ws_summary['A7'] = 'Statistic'
ws_summary['B7'] = 'Value (px)'
ws_summary['C7'] = f'Value (μm, at {UM_PER_PX} μm/px)'
for ci in [1, 2, 3]:
    ws_summary.cell(row=7, column=ci).font = header_font
    ws_summary.cell(row=7, column=ci).fill = header_fill

stats = [
    ('D_eq Min', np.min),
    ('D_eq Max', np.max),
    ('D_eq Mean', np.mean),
    ('D_eq Median', np.median),
    ('D_eq Std', np.std),
]
for i, (label, func) in enumerate(stats, 8):
    ws_summary.cell(row=i, column=1, value=label)
    ws_summary.cell(row=i, column=2, value=round(func(d_eq_px_arr), 2))
    ws_summary.cell(row=i, column=3, value=round(func(d_eq_um_arr), 1))

d32_px = sum(d_eq_px_arr**3) / sum(d_eq_px_arr**2) if sum(d_eq_px_arr**2) > 0 else 0
d32_um = d32_px * UM_PER_PX
r = len(stats) + 8
ws_summary.cell(row=r, column=1, value='D32 (Sauter)')
ws_summary.cell(row=r, column=2, value=round(d32_px, 2))
ws_summary.cell(row=r, column=3, value=round(d32_um, 1))

# Binned distribution
r += 2
ws_summary.cell(row=r, column=1, value='Size Distribution Bins').font = Font(bold=True)
r += 1
ws_summary.cell(row=r, column=1, value='Bin (μm)'); ws_summary.cell(row=r, column=2, value='Count'); ws_summary.cell(row=r, column=3, value='Percentage')
for ci in [1,2,3]:
    ws_summary.cell(row=r, column=ci).font = header_font
    ws_summary.cell(row=r, column=ci).fill = header_fill

bins_um = [0, 200, 400, 600, 800, 1000, 1500, 2000, 3000, 5000, 10000]
for bi in range(len(bins_um) - 1):
    lo, hi = bins_um[bi], bins_um[bi+1]
    n = int(((d_eq_um_arr >= lo) & (d_eq_um_arr < hi)).sum())
    r += 1
    ws_summary.cell(row=r, column=1, value=f'{lo} – {hi}')
    ws_summary.cell(row=r, column=2, value=n)
    ws_summary.cell(row=r, column=3, value=round(n / len(d_eq_um_arr) * 100, 2))

# All droplets sheet
ws_detail = wb.create_sheet('All Droplets')
headers = ['Image', 'Droplet ID',
           'Center X (px)', 'Center Y (px)',
           'Center X (μm)', 'Center Y (μm)',
           'Area (px²)', 'Area (μm²)',
           'Eq Diameter (px)', 'Eq Diameter (μm)',
           'Major Axis (px)', 'Major Axis (μm)',
           'Minor Axis (px)', 'Minor Axis (μm)',
           'Axis Ratio', 'Angle (°)']
for ci, h in enumerate(headers, 1):
    cell = ws_detail.cell(row=1, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill

for ri, d in enumerate(tqdm(all_droplets, desc='Writing Excel'), 2):
    ws_detail.cell(row=ri, column=1, value=d['image'])
    ws_detail.cell(row=ri, column=2, value=d['droplet_id'])
    ws_detail.cell(row=ri, column=3, value=round(d['center_x_px'], 1))
    ws_detail.cell(row=ri, column=4, value=round(d['center_y_px'], 1))
    ws_detail.cell(row=ri, column=5, value=d['center_x_um'])
    ws_detail.cell(row=ri, column=6, value=d['center_y_um'])
    ws_detail.cell(row=ri, column=7, value=d['area_px'])
    ws_detail.cell(row=ri, column=8, value=d['area_um2'])
    ws_detail.cell(row=ri, column=9, value=d['eq_diameter_px'])
    ws_detail.cell(row=ri, column=10, value=d['eq_diameter_um'])
    ws_detail.cell(row=ri, column=11, value=d['major_axis_px'])
    ws_detail.cell(row=ri, column=12, value=d['major_axis_um'])
    ws_detail.cell(row=ri, column=13, value=d['minor_axis_px'])
    ws_detail.cell(row=ri, column=14, value=d['minor_axis_um'])
    ws_detail.cell(row=ri, column=15, value=d['axis_ratio'])
    ws_detail.cell(row=ri, column=16, value=d['angle_deg'])

excel_path = OUT_DIR / 'droplet_statistics.xlsx'
wb.save(str(excel_path))
print(f'Excel saved: {excel_path}')

# ===== Size Distribution Plots =====
fig, axes = plt.subplots(2, 2, figsize=(16, 11))
fig.suptitle(f'Droplet Size Distribution — {FR}\nScale: {UM_PER_PX} μm/px | {len(all_d_eq_px):,} droplets, {len(jfiles)} images',
             fontsize=13, fontweight='bold')

# 1. Histogram — D_eq (px axis with μm twin)
ax1 = axes[0, 0]
ax1.hist(d_eq_px_arr, bins=80, color='steelblue', edgecolor='white', alpha=0.8)
ax1.axvline(np.mean(d_eq_px_arr), color='red', linestyle='--', linewidth=1.5,
            label=f'Mean = {np.mean(d_eq_um_arr):.0f} μm')
ax1.axvline(np.median(d_eq_px_arr), color='orange', linestyle='--', linewidth=1.5,
            label=f'Median = {np.median(d_eq_um_arr):.0f} μm')
ax1.set_xlabel('Equivalent Diameter (px)')
ax1.set_ylabel('Count')
ax1.set_title('D_eq Histogram')
ax1.legend(fontsize=8)
ax1_um = ax1.twiny()
ax1_um.set_xlabel('Equivalent Diameter (μm)', fontsize=9)
ax1_um.set_xlim(ax1.get_xlim()[0] * UM_PER_PX, ax1.get_xlim()[1] * UM_PER_PX)

# 2. Histogram — D_eq in μm
ax2 = axes[0, 1]
ax2.hist(d_eq_um_arr, bins=80, color='darkgreen', edgecolor='white', alpha=0.8)
ax2.axvline(np.mean(d_eq_um_arr), color='red', linestyle='--', linewidth=1.5,
            label=f'Mean = {np.mean(d_eq_um_arr):.0f} μm')
ax2.axvline(np.median(d_eq_um_arr), color='orange', linestyle='--', linewidth=1.5,
            label=f'Median = {np.median(d_eq_um_arr):.0f} μm')
ax2.set_xlabel('Equivalent Diameter (μm)')
ax2.set_ylabel('Count')
ax2.set_title('D_eq Histogram (μm)')
ax2.legend(fontsize=8)

# 3. Cumulative distribution (μm)
ax3 = axes[1, 0]
sorted_d = np.sort(d_eq_um_arr)
cumulative = np.arange(1, len(sorted_d) + 1) / len(sorted_d) * 100
ax3.plot(sorted_d, cumulative, color='darkred', linewidth=2)
ax3.axhline(50, color='gray', linestyle=':', alpha=0.5)
ax3.axvline(np.median(d_eq_um_arr), color='gray', linestyle=':', alpha=0.5,
            label=f'D50 = {np.median(d_eq_um_arr):.0f} μm')
# D10, D90
d10 = np.percentile(d_eq_um_arr, 10); d90 = np.percentile(d_eq_um_arr, 90)
ax3.axvline(d10, color='blue', linestyle='--', alpha=0.5, label=f'D10 = {d10:.0f} μm')
ax3.axvline(d90, color='blue', linestyle='--', alpha=0.5, label=f'D90 = {d90:.0f} μm')
ax3.set_xlabel('Equivalent Diameter (μm)')
ax3.set_ylabel('Cumulative %')
ax3.set_title(f'Cumulative Distribution')
ax3.legend(fontsize=8)
ax3.grid(True, alpha=0.3)

# 4. Axis ratio
ax4 = axes[1, 1]
ratios = np.array([d['axis_ratio'] for d in all_droplets])
ax4.hist(ratios.clip(1, 3), bins=50, color='purple', edgecolor='white', alpha=0.8, range=(1, 3))
ax4.axvline(np.mean(ratios), color='red', linestyle='--', linewidth=1.5,
            label=f'Mean = {np.mean(ratios):.2f}')
ax4.set_xlabel('Major / Minor Axis Ratio')
ax4.set_ylabel('Count')
ax4.set_title('Ellipse Axis Ratio Distribution')
ax4.legend(fontsize=8)

plt.tight_layout()
plot_path = OUT_DIR / 'size_distribution.png'
fig.savefig(str(plot_path), dpi=150, bbox_inches='tight')
plt.close()
print(f'Plot saved: {plot_path}')

# ===== Per-image CSV (with μm) =====
csv_dir = OUT_DIR / 'per_image_csv'
csv_dir.mkdir(exist_ok=True)
by_image = defaultdict(list)
for d in all_droplets:
    by_image[d['image']].append(d)

for img_name, droplets in tqdm(sorted(by_image.items()), desc='Per-image CSV'):
    csv_path = csv_dir / img_name.replace('.bmp', '.csv')
    with open(csv_path, 'w') as f:
        f.write(f'# Scale: {UM_PER_PX} um/px\n')
        f.write('droplet_id,center_x_px,center_y_px,center_x_um,center_y_um,area_px,area_um2,eq_diameter_px,eq_diameter_um,major_axis_px,major_axis_um,minor_axis_px,minor_axis_um,axis_ratio,angle_deg\n')
        for d in droplets:
            f.write(f"{d['droplet_id']},{d['center_x_px']:.1f},{d['center_y_px']:.1f},{d['center_x_um']},{d['center_y_um']},{d['area_px']},{d['area_um2']},{d['eq_diameter_px']},{d['eq_diameter_um']},{d['major_axis_px']},{d['major_axis_um']},{d['minor_axis_px']},{d['minor_axis_um']},{d['axis_ratio']},{d['angle_deg']}\n")

print(f'Per-image CSVs: {csv_dir} ({len(by_image)} files)')
print(f'\nAll outputs saved to: {OUT_DIR}')
print(f'  Scale: {UM_PER_PX} μm/px')
print(f'  {len(jfiles)} annotated images -> output/annotated/{FR}/')
print(f'  Excel: output/results/{FR}/droplet_statistics.xlsx')
print(f'  Plots: output/results/{FR}/size_distribution.png')
print(f'  Per-image CSV: output/results/{FR}/per_image_csv/')
