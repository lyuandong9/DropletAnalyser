"""
ImageIdentification — Desktop Application
Complete PyQt6 app: crop + detect + review + export
"""

import sys, os, json, math, traceback, datetime, multiprocessing
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import cv2, numpy as np
from skimage.feature import peak_local_max
from scipy import ndimage as ndi
from skimage.segmentation import watershed

from PyQt6.QtWidgets import *
from PyQt6.QtGui import *
from PyQt6.QtCore import *

# Lazy import — matplotlib backend must be set AFTER QApplication exists in frozen apps
FigureCanvasQTAgg = None
Figure = None
Workbook = None
XlFont = None
PatternFill = None

def _lazy_import():
    global FigureCanvasQTAgg, Figure, Workbook, XlFont, PatternFill
    if FigureCanvasQTAgg is None:
        import matplotlib
        matplotlib.use('QtAgg')
        from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as _fc
        from matplotlib.figure import Figure as _f
        FigureCanvasQTAgg = _fc
        Figure = _f
        from openpyxl import Workbook as _wb
        Workbook = _wb
        from openpyxl.styles import Font as _xf, PatternFill as _pf
        XlFont = _xf
        PatternFill = _pf

UM_PER_PX_DEFAULT = 110.0



def _imread(path):
    """Read image from any path, handling non-ASCII filenames on Windows."""
    img = cv2.imread(str(path))
    if img is None:
        with open(str(path), 'rb') as f:
            data = f.read()
        img = cv2.imdecode(np.frombuffer(data, np.uint8), cv2.IMREAD_COLOR)
    return img



# ── Detection Engine ───────────────────────────────────────────

def detect_droplets(gray, bg_sigma=80, min_distance=5, peak_threshold=0.04,
                    seed_threshold=0.05, min_area=10, max_area=350, max_aspect=2.0,
                    scale_factor=1.0, area_threshold=0.35, area_penalty=250,
                    edge_margin=5, median_bg=None):
    h, w = gray.shape

    if median_bg is not None:
        # Auto-detect droplet polarity: droplets can be brighter OR darker than median bg
        # Use the direction with more binary pixels above threshold (not max value)
        residual_a = cv2.subtract(median_bg, gray)  # dark droplets on bright bg
        residual_b = cv2.subtract(gray, median_bg)  # bright droplets on dark bg
        if (residual_b > 6).sum() > (residual_a > 6).sum():
            residual = residual_b
        else:
            residual = residual_a
        binary = residual > 6
        k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
        binary = cv2.morphologyEx(binary.astype(np.uint8), cv2.MORPH_OPEN, k, iterations=1)
        binary = binary.astype(bool)
        dist_map = ndi.distance_transform_edt(binary)
        coords = peak_local_max(dist_map, min_distance=9, threshold_rel=0.04, exclude_border=edge_margin)
        if len(coords) == 0:
            return []
        markers = np.zeros((h, w), dtype=np.int32)
        for i, (r, c) in enumerate(coords):
            markers[r, c] = i + 1
        labels = watershed(-dist_map, markers, mask=binary)
        dets = []
        for lid in range(1, labels.max() + 1):
            region = labels == lid
            n_px = int(region.sum())
            if n_px < 6 or n_px > max_area:
                continue
            ys, xs = np.where(region)
            if len(ys) < 5:
                continue
            if (ys.min() <= edge_margin or xs.min() <= edge_margin or
                ys.max() >= h - edge_margin - 1 or xs.max() >= w - edge_margin - 1):
                continue
            d_eq_px = round(2 * math.sqrt(n_px / math.pi), 1)
            vis_r = max(1, int(d_eq_px * scale_factor / 2))
            cx = xs.mean(); cy = ys.mean()
            pts = cv2.ellipse2Poly((int(cx), int(cy)), (vis_r, vis_r), 0, 0, 360, 32)
            dets.append({'id': len(dets), 'contour': pts.tolist(),
                         'center': [round(float(cx), 1), round(float(cy), 1)],
                         'axes': [d_eq_px, d_eq_px],
                         'angle': 0.0, 'area_px': n_px, 'type': 'ellipse'})
        dets.sort(key=lambda d: d['area_px'], reverse=True)
        return dets

    # Gaussian path: original generic algorithm for single-frame images
    bg = cv2.GaussianBlur(gray, (251, 251), 80)
    residual = cv2.subtract(bg, gray)
    normalized = residual.astype(float) / np.clip(bg, 1, 255)
    bs = normalized > normalized.max() * seed_threshold
    k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
    bs = cv2.morphologyEx(bs.astype(np.uint8), cv2.MORPH_OPEN, k, iterations=1).astype(bool)
    coords = peak_local_max(ndi.distance_transform_edt(bs),
                            min_distance=min_distance, threshold_rel=peak_threshold,
                            exclude_border=edge_margin)
    if len(coords) == 0:
        return []
    markers = np.zeros((h, w), dtype=np.int32)
    for i, (r, c) in enumerate(coords):
        markers[r, c] = i + 1
    labels = watershed(-normalized.astype(float), markers, mask=bs)
    dets = []
    for lid in range(1, labels.max() + 1):
        region = labels == lid
        n_px = int(region.sum())
        if n_px < min_area or n_px > max_area:
            continue
        ys, xs = np.where(region)
        if len(ys) < 5:
            continue
        if (ys.min() <= edge_margin or xs.min() <= edge_margin or
            ys.max() >= h - edge_margin - 1 or xs.max() >= w - edge_margin - 1):
            continue
        if n_px > area_penalty:
            bw = region.astype(np.uint8)
            cnts, _ = cv2.findContours(bw, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            if cnts:
                if n_px / cv2.contourArea(cv2.convexHull(cnts[0])) < 0.7:
                    continue
        rv = residual[region].astype(np.float64)
        pk = rv.max()
        if pk < 1:
            continue
        eff_a = (rv >= pk * area_threshold).sum()
        if eff_a < min_area:
            continue
        ww = residual[region].astype(np.float64)
        tw = ww.sum()
        if tw == 0:
            continue
        cx = (xs * ww).sum() / tw; cy = (ys * ww).sum() / tw
        er = math.sqrt(eff_a / math.pi)
        if (cx - er < edge_margin or cy - er < edge_margin or
            cx + er > w - edge_margin or cy + er > h - edge_margin):
            continue
        dx, dy = xs - cx, ys - cy
        mu20 = (dx*dx*ww).sum()/tw; mu02 = (dy*dy*ww).sum()/tw; mu11 = (dx*dy*ww).sum()/tw
        evals, evecs = np.linalg.eigh([[mu20, mu11], [mu11, mu02]])
        maj_raw = 2 * math.sqrt(max(evals[0], evals[1], 0.01))
        mn_raw = 2 * math.sqrt(max(min(evals[0], evals[1]), 0.01))
        if maj_raw / max(mn_raw, 0.1) > max_aspect:
            continue
        d_eq_px = round(2 * math.sqrt(eff_a / math.pi), 1)
        vis_r = int(d_eq_px * scale_factor / 2)
        ang = math.degrees(math.atan2(evecs[1, 0], evecs[0, 0]))
        pts = cv2.ellipse2Poly((int(cx), int(cy)), (vis_r, vis_r), int(ang), 0, 360, 32)
        dets.append({'id': len(dets), 'contour': pts.tolist(),
                     'center': [round(float(cx), 1), round(float(cy), 1)],
                     'axes': [d_eq_px, d_eq_px],
                     'angle': round(float(ang), 1), 'area_px': int(eff_a), 'type': 'ellipse'})
    dets.sort(key=lambda d: d['area_px'], reverse=True)
    return dets

# ── Workers ────────────────────────────────────────────────────

def _build_median_bg(paths):
    """Build temporal median background from ALL frames."""
    if len(paths) < 3:
        return None
    stack = []
    for p in sorted(paths):
        img = _imread(p)
        if img is not None:
            if len(img.shape)==3:
                img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            stack.append(img)
    if len(stack) < 3:
        return None
    return np.median(np.array(stack), axis=0).astype(np.uint8)


class DetectWorker(QThread):
    progress = pyqtSignal(int,int)
    finished = pyqtSignal(dict)
    status = pyqtSignal(str)
    error = pyqtSignal(str)
    def __init__(self, paths, params, num_workers=None, use_temporal=False):
        super().__init__()
        self.paths=list(paths); self.params=params
        if num_workers is None or num_workers <= 0:
            if getattr(sys, 'frozen', False):
                self.nw = 1  # frozen app: avoid multiprocessing deadlocks
            else:
                self.nw = max(1, multiprocessing.cpu_count())
        else:
            self.nw = num_workers
        self.use_temporal = use_temporal
        self.median_bg = None
    def run(self):
        # Build temporal median background if requested
        if self.use_temporal:
            self.status.emit('Building temporal median background...')
            self.median_bg = _build_median_bg(self.paths)
            if self.median_bg is not None:
                self.params['median_bg'] = self.median_bg
            else:
                self.params.pop('median_bg', None)

        result={}; total=len(self.paths)
        if (self.nw <= 1 or total <= 1) and self.median_bg is None:
            for i,p in enumerate(self.paths):
                self.status.emit(f'Detecting: {Path(p).name}')
                try:
                    img=_imread(p)
                    if img is not None:
                        result[Path(p).name]=detect_droplets(cv2.cvtColor(img,cv2.COLOR_BGR2GRAY),**self.params)
                except Exception as e:
                    self.error.emit(f'Detection error: {e}')
                self.progress.emit(i+1,total)
            self.finished.emit(result); return

        self.status.emit(f'Detecting using {self.nw} cores...')
        done=0
        params = dict(self.params)  # copy so threads don't race on median_bg
        median_bg = params.pop('median_bg', None)  # extract — passed separately to each thread
        def _process_one(p):
            try:
                img = _imread(p)
                if img is None: return Path(p).name, []
                gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                if median_bg is not None:
                    d = detect_droplets(gray, median_bg=median_bg, **params)
                else:
                    d = detect_droplets(gray, **params)
                return Path(p).name, d
            except Exception as e2:
                return Path(p).name, []
        with ThreadPoolExecutor(max_workers=self.nw) as ex:
            futures = {ex.submit(_process_one, p): p for p in self.paths}
            for f in as_completed(futures):
                try:
                    name,d=f.result()
                    result[name]=d
                except Exception as e:
                    self.error.emit(f'Error: {futures[f].name}: {e}')
                done+=1; self.progress.emit(done,total)
        self.finished.emit(result)

class ExportWorker(QThread):
    progress = pyqtSignal(int,int)
    finished = pyqtSignal(str)
    status = pyqtSignal(str)
    def __init__(self,src_dir,det_dir,out_dir,um):
        super().__init__(); self.src=Path(src_dir); self.det=Path(det_dir)
        self.out=Path(out_dir); self.um=um
    def run(self):
        self.out.mkdir(parents=True,exist_ok=True)
        (self.out/'annotated').mkdir(exist_ok=True)
        jfs=sorted(self.det.glob('*.json'))
        if not jfs: self.finished.emit('No files'); return
        rows=[]; cols=[(0,255,0),(255,255,0),(0,255,255),(255,0,255),(0,128,255),(255,128,0),(128,255,0),(0,0,255)]
        for i,jf in enumerate(jfs):
            self.status.emit(f'Exporting: {jf.stem}')
            bmp=self.src/(jf.stem+'.bmp')
            img=_imread(bmp)
            with open(jf) as f: dets=json.load(f)
            if img is not None:
                for d in dets.get('droplets',[]):
                    cv2.polylines(img,[np.array(d['contour'],np.int32)],True,cols[d['id']%8],1)
                cv2.imwrite(str(self.out/'annotated'/(jf.stem+'.jpg')),img,[cv2.IMWRITE_JPEG_QUALITY,88])
            for d in dets.get('droplets',[]):
                a=d['area_px']; de=2*math.sqrt(a/math.pi); maj,mn=d['axes']
                rows.append({'image':bmp.name,'id':d['id'],'cx_px':d['center'][0],'cy_px':d['center'][1],
                    'cx_um':d['center'][0]*self.um,'cy_um':d['center'][1]*self.um,
                    'area_px':a,'area_um2':a*self.um*self.um,'d_eq_px':de,'d_eq_um':de*self.um,
                    'maj_px':maj,'maj_um':maj*self.um,'min_px':mn,'min_um':mn*self.um,
                    'ratio':maj/max(mn,0.1),'angle':d['angle']})
            self.progress.emit(i+1,len(jfs))
        vals=np.array([r['d_eq_um'] for r in rows])
        _lazy_import()
        wb=Workbook(); ws=wb.active; ws.title='Summary'
        ws['A1']='Droplet Analysis Report'; ws['A2']='Scale (um/px)'; ws['B2']=self.um
        ws['A3']='Total Images'; ws['B3']=len(jfs); ws['A4']='Total Droplets'; ws['B4']=len(vals)
        ws['A6']='Statistic'; ws['B6']='Value (um)'
        for ri,(l,f) in enumerate([('Min',np.min),('Max',np.max),('Mean',np.mean),('Median',np.median),('Std',np.std)],7):
            ws.cell(row=ri,column=1,value=l); ws.cell(row=ri,column=2,value=round(f(vals),1))
        d32=sum(vals**3)/sum(vals**2) if sum(vals**2)>0 else 0
        ws.cell(row=12,column=1,value='D32 (Sauter)'); ws.cell(row=12,column=2,value=round(d32,1))
        ws2=wb.create_sheet('All Droplets')
        hds=['Image','ID','CenterX(px)','CenterY(px)','CenterX(um)','CenterY(um)','Area(px2)','Area(um2)',
             'D_eq(px)','D_eq(um)','Maj(px)','Maj(um)','Min(px)','Min(um)','Ratio','Angle']
        hf=XlFont(bold=True,color='FFFFFF'); hfl=PatternFill('solid',fgColor='4472C4')
        for ci,h in enumerate(hds,1): c=ws2.cell(row=1,column=ci,value=h); c.font=hf; c.fill=hfl
        keys=['image','id','cx_px','cy_px','cx_um','cy_um','area_px','area_um2','d_eq_px','d_eq_um',
              'maj_px','maj_um','min_px','min_um','ratio','angle']
        for ri,r in enumerate(rows,2):
            for ci,k in enumerate(keys,1):
                ws2.cell(row=ri,column=ci,value=round(r[k],2) if isinstance(r[k],float) else r[k])
        wb.save(str(self.out/'droplet_statistics.xlsx'))
        fig=Figure(figsize=(12,8)); axes=fig.subplots(2,2)
        ax=axes[0,0]; ax.hist(vals,bins=80,color='steelblue',edgecolor='white',alpha=0.8)
        ax.axvline(np.mean(vals),color='red',ls='--',label=f'Mean={np.mean(vals):.0f}um')
        ax.set_xlabel('D_eq (um)'); ax.set_ylabel('Count'); ax.legend(fontsize=7)
        ax=axes[0,1]; lb=np.logspace(np.log10(max(1,vals.min())),np.log10(vals.max()),60)
        ax.hist(vals,bins=lb,color='darkgreen',edgecolor='white',alpha=0.8)
        ax.set_xscale('log'); ax.set_xlabel('D_eq (um,log)'); ax.set_ylabel('Count')
        ax=axes[1,0]; sd=np.sort(vals); ax.plot(sd,np.arange(1,len(sd)+1)/len(sd)*100,color='darkred',lw=2)
        for dv,dl in [(np.median(vals),'D50'),(np.percentile(vals,10),'D10'),(np.percentile(vals,90),'D90')]:
            ax.axvline(dv,ls='--',alpha=0.5,label=f'{dl}={dv:.0f}um')
        ax.set_xlabel('D_eq (um)'); ax.set_ylabel('Cumulative %'); ax.legend(fontsize=7); ax.grid(alpha=0.3)
        ax=axes[1,1]; ratios=np.array([r['ratio'] for r in rows])
        ax.hist(np.clip(ratios,1,3),bins=50,color='purple',edgecolor='white',alpha=0.8,range=(1,3))
        ax.set_xlabel('Axis Ratio'); ax.set_ylabel('Count')
        fig.tight_layout(); fig.savefig(str(self.out/'size_distribution.png'),dpi=150)
        self.finished.emit(str(self.out))

# ── Image Viewer ──────────────────────────────────────────────

class ImageViewer(QGraphicsView):
    dropletSelected = pyqtSignal(int)
    def __init__(self):
        super().__init__()
        self._scene=QGraphicsScene(); self.setScene(self._scene)
        self.setRenderHint(QPainter.RenderHint.Antialiasing)
        self.setTransformationAnchor(QGraphicsView.ViewportAnchor.AnchorUnderMouse)
        self.setResizeAnchor(QGraphicsView.ViewportAnchor.AnchorUnderMouse)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.setDragMode(QGraphicsView.DragMode.ScrollHandDrag)
        self._px=None; self._droplets=[]; self._sel=-1
        self._cols=[QColor(0,255,0),QColor(255,255,0),QColor(0,255,255),QColor(255,0,255),
                    QColor(0,128,255),QColor(255,128,0),QColor(128,255,0),QColor(0,0,255)]
        self._mode='view'; self._add_center=None; self._add_radius=0
        self._next_id=0; self._pan_start=None; self._modified=False
        self._current_gray=None; self._hover_droplet=None
        self.viewport().installEventFilter(self)

    def eventFilter(self, obj, event):
        if obj!=self.viewport(): return super().eventFilter(obj,event)
        from PyQt6.QtGui import QMouseEvent
        t=event.type()
        if isinstance(event, QMouseEvent):
            pt=event.position().toPoint()
            if t==QEvent.Type.MouseButtonPress and event.button()==Qt.MouseButton.LeftButton:
                return self._on_press(pt)
            elif t==QEvent.Type.MouseMove:
                return self._on_move(pt)
            elif t==QEvent.Type.MouseButtonRelease and event.button()==Qt.MouseButton.LeftButton:
                return self._on_release(pt)
        return super().eventFilter(obj, event)

    def _on_press(self, pos):
        try:
            sp=self.mapToScene(pos); mx,my=sp.x(),sp.y()
            if self._mode=='add':
                if self._hover_droplet is not None:
                    droplet = dict(self._hover_droplet)
                    droplet['id'] = self._next_id; self._next_id += 1
                    droplet['manual'] = True; self._modified = True
                    self._droplets.append(droplet)
                    self._sel = len(self._droplets) - 1
                    self._hover_droplet = None
                    self._redraw(); self.dropletSelected.emit(self._sel)
                else:
                    # Fall back to auto-detect, then freehand if nothing found
                    droplet = self._auto_detect_at(int(mx), int(my))
                    if droplet:
                        droplet['id'] = self._next_id; self._next_id += 1
                        droplet['manual'] = True; self._modified = True
                        self._droplets.append(droplet)
                        self._sel = len(self._droplets) - 1
                        self._redraw(); self.dropletSelected.emit(self._sel)
                    else:
                        self._add_center = (mx, my)
                        self._draw_pts = [(mx, my)]
                return True
            if self._mode=='delete':
                best,bd=-1,12
                for i,d in enumerate(self._droplets):
                    for pt in d.get('contour',[]):
                        dd=((pt[0]-mx)**2+(pt[1]-my)**2)**0.5
                        if dd<bd: bd=dd; best=i
                if best>=0:
                    del self._droplets[best]; self._modified=True
                    if self._sel==best: self._sel=-1
                    self._redraw(); self.dropletSelected.emit(-1)
                return True
            self._pan_start=pos
        except Exception:
            pass
        return False

    def _on_move(self, pos):
        try:
            if self._mode=='add' and self._add_center is not None:
                sp=self.mapToScene(pos); mx,my=sp.x(),sp.y()
                if self._draw_pts:
                    lx,ly=self._draw_pts[-1]
                    if ((mx-lx)**2+(my-ly)**2)**0.5>=4:
                        steps=max(1,int(((mx-lx)**2+(my-ly)**2)**0.5/3))
                        for s in range(1,steps+1):
                            t=s/(steps+1)
                            self._draw_pts.append((lx+(mx-lx)*t, ly+(my-ly)*t))
                else:
                    self._draw_pts.append((mx,my))
                self._redraw(); return True
            elif self._mode=='add':
                sp=self.mapToScene(pos); mx,my=sp.x(),sp.y()
                droplet=self._auto_detect_at(int(mx),int(my))
                if droplet:
                    # Compare centers to avoid flickering redraws
                    prev=self._hover_droplet
                    if (prev is None or
                        abs(prev['center'][0]-droplet['center'][0])>1 or
                        abs(prev['center'][1]-droplet['center'][1])>1):
                        self._hover_droplet=droplet; self._redraw()
                else:
                    if self._hover_droplet is not None:
                        self._hover_droplet=None; self._redraw()
                return True
        except Exception:
            pass
        return False

    def _on_release(self, pos):
        try:
            if self._mode=='add' and self._add_center is not None:
                pts=self._draw_pts
                if len(pts)>=3:
                    area=0.0; n=len(pts)
                    for i in range(n):
                        j=(i+1)%n
                        area+=pts[i][0]*pts[j][1]-pts[j][0]*pts[i][1]
                    area_px=abs(area)/2.0
                    if area_px>=5:
                        d_eq=2*math.sqrt(area_px/math.pi)
                        cx=sum(p[0] for p in pts)/n; cy=sum(p[1] for p in pts)/n
                        contour=[[p[0],p[1]] for p in pts]
                        self._droplets.append({
                            'id':self._next_id,
                            'contour':contour,
                            'center':[round(cx,1),round(cy,1)],
                            'axes':[d_eq,d_eq],'angle':0,
                            'area_px':int(area_px),'type':'polygon','manual':True})
                        self._next_id+=1; self._modified=True
                self._add_center=None; self._draw_pts=[]
                self._redraw(); self.dropletSelected.emit(-1); return True
            if self._mode=='view' and self._pan_start:
                d2=(pos.x()-self._pan_start.x())**2+(pos.y()-self._pan_start.y())**2
                if d2<25:
                    sp=self.mapToScene(pos)
                    self._select(sp.x(),sp.y())
                self._pan_start=None
        except Exception:
            pass
        return False

    def set_median_bg(self, mb):
        self._median_bg = mb

    def _auto_detect_at(self, mx, my):
        """Auto-detect droplet at clicked point — uses median bg if available."""
        if self._current_gray is None:
            return None
        gray = self._current_gray; h,w = gray.shape
        r = 25
        r1,r2 = max(0,my-r), min(h,my+r+1)
        c1,c2 = max(0,mx-r), min(w,mx+r+1)
        local = gray[r1:r2, c1:c2]
        if local.size < 25:
            return None

        mb = getattr(self, '_median_bg', None)
        if mb is not None:
            # Use the global median background for this local patch
            local_mb = mb[r1:r2, c1:c2]
            res_a = cv2.subtract(local_mb, local)  # dark droplets
            res_b = cv2.subtract(local, local_mb)  # bright droplets
            res = res_b if res_b.max() > res_a.max() else res_a
            binary_patch = res > 5
            k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3,3))
            binary_patch = cv2.morphologyEx(binary_patch.astype(np.uint8), cv2.MORPH_OPEN, k, iterations=1).astype(bool)
        else:
            bg = cv2.GaussianBlur(local, (51,51), 30)
            res = cv2.subtract(bg, local)
            res_norm = res.astype(float) / np.clip(bg, 1, 255)
            binary_patch = res_norm > res_norm.max() * 0.05
            k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3,3))
            binary_patch = cv2.morphologyEx(binary_patch.astype(np.uint8), cv2.MORPH_OPEN, k, iterations=1).astype(bool)

        lr,lc = my-r1, mx-c1
        labeled, n = ndi.label(binary_patch)
        if n == 0 or labeled[lr,lc] == 0:
            return None
        region = labeled == labeled[lr,lc]
        area = int(region.sum())
        if area < 5 or area > 500:
            return None
        ys,xs = np.where(region)
        if len(ys) < 5:
            return None
        gxs = xs + c1; gys = ys + r1
        cx = gxs.mean(); cy = gys.mean()
        d_eq = round(2 * math.sqrt(area / math.pi), 1)
        r_vis = max(1, int(d_eq  / 2))
        pts = cv2.ellipse2Poly((int(cx), int(cy)), (r_vis, r_vis), 0, 0, 360, 32)
        return {'contour': pts.tolist(), 'center': [round(float(cx),1), round(float(cy),1)],
                'axes': [d_eq, d_eq], 'angle': 0, 'area_px': area, 'type': 'ellipse'}

    def _select(self, x, y):
        best,bd=-1,12
        for i,d in enumerate(self._droplets):
            for p in d.get('contour',[]):
                dd=((p[0]-x)**2+(p[1]-y)**2)**0.5
                if dd<bd: bd=dd; best=i
        self._sel=-1 if self._sel==best else best
        self._redraw()
        self.dropletSelected.emit(self._sel)

    def set_image(self, arr):
        self._scene.clear(); self._px=None
        if arr is None: return
        self._current_gray = arr if len(arr.shape)==2 else cv2.cvtColor(arr, cv2.COLOR_BGR2GRAY)
        if len(arr.shape)==2:
            q=QImage(arr.data,arr.shape[1],arr.shape[0],arr.shape[1],QImage.Format.Format_Grayscale8)
        else:
            rgb=cv2.cvtColor(arr,cv2.COLOR_BGR2RGB)
            q=QImage(rgb.data,arr.shape[1],arr.shape[0],arr.shape[1]*3,QImage.Format.Format_RGB888)
        self._px=self._scene.addPixmap(QPixmap.fromImage(q.copy()))
        self._scene.setSceneRect(QRectF(QPointF(0,0),QSizeF(arr.shape[1],arr.shape[0])))
        self.fitInView(self._scene.sceneRect(),Qt.AspectRatioMode.KeepAspectRatio)

    def set_droplets(self, lst):
        self._droplets=lst or []; self._sel=-1; self._modified=False
        self._next_id=max((d.get('id',0) for d in self._droplets),default=-1)+1 if self._droplets else 0
        self._redraw()

    def _redraw(self):
        for item in list(self._scene.items()):
            if item!=self._px: self._scene.removeItem(item)
        if not self._droplets: return
        for i,d in enumerate(self._droplets):
            sel=i==self._sel
            c=QColor('#ff0') if sel else self._cols[d.get('id',i)%len(self._cols)]
            pen=QPen(c,2.5 if sel else 1.5); pen.setCosmetic(True)
            brush=QBrush(QColor(255,255,0,35) if sel else QColor(0,255,0,15))
            pts=d.get('contour',[])
            if pts and len(pts)>=3:
                poly=QPolygonF(); [poly.append(QPointF(p[0],p[1])) for p in pts]
                self._scene.addPolygon(poly,pen,brush)
        if self._mode=='add' and self._hover_droplet is not None and self._add_center is None:
            d=self._hover_droplet
            pen=QPen(QColor(0,200,220),2); pen.setCosmetic(True); pen.setStyle(Qt.PenStyle.DashLine)
            brush=QBrush(QColor(0,200,220,25))
            pts=d.get('contour',[])
            if pts and len(pts)>=3:
                poly=QPolygonF(); [poly.append(QPointF(p[0],p[1])) for p in pts]
                self._scene.addPolygon(poly,pen,brush)
        if self._mode=='add' and self._add_center and hasattr(self,'_draw_pts') and len(self._draw_pts)>=2:
            pen=QPen(QColor(0,200,0),2); pen.setCosmetic(True); pen.setStyle(Qt.PenStyle.DashLine)
            path=QPainterPath()
            path.moveTo(self._draw_pts[0][0],self._draw_pts[0][1])
            for pt in self._draw_pts[1:]:
                path.lineTo(pt[0],pt[1])
            self._scene.addPath(path,pen)
            for pt in self._draw_pts:
                self._scene.addEllipse(QPointF(pt[0],pt[1]),1,1,QPen(QColor(0,200,0)),QBrush(QColor(0,200,0)))

    def wheelEvent(self, e):
        self.scale(1.15 if e.angleDelta().y()>0 else 1/1.15,1.15 if e.angleDelta().y()>0 else 1/1.15)
        self._redraw()

    def set_mode(self, m):
        self._mode=m; self._add_center=None; self._draw_pts=[]
        self._hover_droplet=None
        if m=='view':
            self.setCursor(QCursor(Qt.CursorShape.ArrowCursor))
            self.setDragMode(QGraphicsView.DragMode.ScrollHandDrag)
        elif m=='add':
            self.setCursor(QCursor(Qt.CursorShape.CrossCursor))
            self.setDragMode(QGraphicsView.DragMode.NoDrag)
        elif m=='delete':
            self.setCursor(QCursor(Qt.CursorShape.ForbiddenCursor))
            self.setDragMode(QGraphicsView.DragMode.NoDrag)
        self._redraw()

    def reset_zoom(self):
        if self._px: self.fitInView(self._scene.sceneRect(),Qt.AspectRatioMode.KeepAspectRatio)
        self._redraw()

    def delete_selected(self):
        if self._sel>=0: del self._droplets[self._sel]; self._sel=-1; self._modified=True; self._redraw()



class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('ImageIdentification — Droplet/Bubble Analyzer')
        self.setMinimumSize(1200,750)
        try:
            self._settings=QSettings('DropletAnalysis','ImageIdentification')
            self._raw_dir=Path(self._settings.value('raw_dir',str(Path.home())))
        except:
            self._settings=QSettings('DropletAnalysis','ImageIdentification')
            self._raw_dir=Path.home()
        self._cropped_dir=None; self._detections={}; self._img_idx=0; self._img_list=[]
        self._crop_rect=None; self._crop_img_w=0; self._crop_img_h=0
        self._setup()
        # Move to visible position — avoid off-screen from saved geometry
        self.move(100, 100)
        try:
            geo=self._settings.value('geometry')
            if geo: self.restoreGeometry(geo)
        except:
            pass

    def _setup(self):
        c=QWidget(); self.setCentralWidget(c); l=QVBoxLayout(c)
        tb=self.addToolBar('Main')
        tb.addAction('Open Images',self._open_raw).setToolTip('Select folder with raw PIV images')
        tb.addAction('Open Cropped',self._open_cropped).setToolTip('Select folder with cropped images')
        self._tabs=QTabWidget(); l.addWidget(self._tabs)
        self._tabs.addTab(self._tab_crop(),'1. Crop Images')
        self._tabs.addTab(self._tab_preprocess(),'2. Preprocess')
        self._tabs.addTab(self._tab_detect(),'3. Detect & Review')
        self._tabs.addTab(self._tab_export(),'4. Export Results')
        self._status=QStatusBar(); self.setStatusBar(self._status)

    # ── Tab 1: Crop ────────────────────────────────────────

    def _tab_crop(self):
        w=QWidget(); l=QVBoxLayout(w)
        cr=QHBoxLayout(); cr.addWidget(QLabel('Source:'))
        self._crop_lbl=QLabel('(none)'); cr.addWidget(self._crop_lbl,1)
        cr.addWidget(QPushButton('Browse...',clicked=self._open_raw))
        sp=QSplitter(Qt.Orientation.Horizontal)
        lt=QWidget(); ll=QVBoxLayout(lt)
        ll.addWidget(QLabel('Subfolders:')); self._crop_flist=QListWidget()
        self._crop_flist.currentItemChanged.connect(self._crop_fchg); ll.addWidget(self._crop_flist)
        ll.addWidget(QLabel('Images:')); self._crop_ilist=QListWidget()
        self._crop_ilist.currentItemChanged.connect(self._crop_ichg); ll.addWidget(self._crop_ilist)
        rt=QWidget(); rl=QVBoxLayout(rt)
        self._crop_view=ImageViewer(); rl.addWidget(self._crop_view)
        ccr=QHBoxLayout(); self._crop_clbl=QLabel('Drag mouse to select crop region')
        ccr.addWidget(self._crop_clbl); ccr.addStretch()
        cc_btn=QPushButton('Apply Crop to This Subfolder')
        cc_btn.setStyleSheet('background:#106040;color:#fff;font-weight:bold;padding:8px')
        cc_btn.clicked.connect(self._apply_crop); ccr.addWidget(cc_btn)
        cc_all_btn=QPushButton('Apply Crop to ALL Sibling Folders')
        cc_all_btn.setStyleSheet('background:#602080;color:#fff;font-weight:bold;padding:8px')
        cc_all_btn.clicked.connect(self._apply_crop_all); ccr.addWidget(cc_all_btn)
        cc_cur_btn=QPushButton('Crop Current Image Only')
        cc_cur_btn.setStyleSheet('background:#806020;color:#fff;font-weight:bold;padding:8px')
        cc_cur_btn.clicked.connect(self._apply_crop_current); ccr.addWidget(cc_cur_btn)
        rl.addLayout(ccr)

        # Save settings row
        # Save settings row — two lines
        save_line1 = QHBoxLayout()
        save_line1.addWidget(QLabel('Save to:'))
        saved_out = self._settings.value('crop_out_dir', '')
        default_out = saved_out if saved_out else str(Path(__file__).parent.parent/'output'/'cropped')
        self._crop_out_dir = QLineEdit(default_out)
        self._crop_out_dir.setReadOnly(True)
        self._crop_out_dir.setStyleSheet('color:#2060a0;font-size:11px')
        save_line1.addWidget(self._crop_out_dir, 1)
        out_btn = QPushButton('Browse...')
        out_btn.setFixedWidth(65)
        out_btn.clicked.connect(self._choose_crop_out); save_line1.addWidget(out_btn)
        rl.addLayout(save_line1)

        # Manual coordinate input row
        coord_line = QHBoxLayout()
        coord_line.addWidget(QLabel('X1:')); self._crop_x1 = QSpinBox(); self._crop_x1.setRange(0,9999); self._crop_x1.setFixedWidth(55)
        coord_line.addWidget(self._crop_x1)
        coord_line.addWidget(QLabel('Y1:')); self._crop_y1 = QSpinBox(); self._crop_y1.setRange(0,9999); self._crop_y1.setFixedWidth(55)
        coord_line.addWidget(self._crop_y1)
        coord_line.addWidget(QLabel('X2:')); self._crop_x2 = QSpinBox(); self._crop_x2.setRange(0,9999); self._crop_x2.setFixedWidth(55)
        coord_line.addWidget(self._crop_x2)
        coord_line.addWidget(QLabel('Y2:')); self._crop_y2 = QSpinBox(); self._crop_y2.setRange(0,9999); self._crop_y2.setFixedWidth(55)
        coord_line.addWidget(self._crop_y2)
        set_btn = QPushButton('Set From Manual')
        set_btn.clicked.connect(self._set_crop_from_manual); coord_line.addWidget(set_btn)
        reuse_btn = QPushButton('Reuse Last Crop')
        reuse_btn.clicked.connect(self._reuse_last_crop); coord_line.addWidget(reuse_btn)
        rl.addLayout(coord_line)

        save_line2 = QHBoxLayout()
        save_line2.addWidget(QLabel('Prefix:'))
        self._crop_prefix = QLineEdit('frame_')
        self._crop_prefix.setToolTip('Filename prefix')
        save_line2.addWidget(self._crop_prefix, 1)

        save_line2.addWidget(QLabel(' Digits:'))
        self._crop_digits = QSpinBox()
        self._crop_digits.setRange(2, 8); self._crop_digits.setValue(4)
        self._crop_digits.setFixedWidth(50)
        self._crop_digits.setToolTip('Zero-padded digits, 4 = 0001, 0002...')
        save_line2.addWidget(self._crop_digits)

        save_line2.addWidget(QLabel(' Start:'))
        self._crop_start = QSpinBox()
        self._crop_start.setRange(0, 99999); self._crop_start.setValue(0)
        self._crop_start.setFixedWidth(65)
        self._crop_start.setToolTip('Starting index for numbering')
        save_line2.addWidget(self._crop_start)
        rl.addLayout(save_line2)
        sp.addWidget(lt); sp.addWidget(rt); sp.setSizes([280,1000])
        l.addLayout(cr); l.addWidget(sp)
        # Rubber band for crop selection
        self._rubber=QRubberBand(QRubberBand.Shape.Rectangle,self._crop_view)
        self._crop_origin=None
        self._crop_view.viewport().installEventFilter(self)
        return w

    def eventFilter(self, obj, event):
        if obj==self._crop_view.viewport():
            tp=event.type()
            if tp==QEvent.Type.MouseButtonPress and event.button()==Qt.MouseButton.LeftButton:
                self._crop_origin=event.position().toPoint()
                self._rubber.setGeometry(QRect(self._crop_origin,QSize()))
                self._rubber.show(); return True
            elif tp==QEvent.Type.MouseMove and self._rubber.isVisible():
                self._rubber.setGeometry(QRect(self._crop_origin,event.position().toPoint()).normalized())
                return True
            elif tp==QEvent.Type.MouseButtonRelease and self._rubber.isVisible():
                self._rubber.hide()
                r=QRect(self._crop_origin,event.position().toPoint()).normalized()
                if r.width()>10 and r.height()>10:
                    sr=self._crop_view.mapToScene(r).boundingRect()
                    self._crop_rect=(int(max(0,sr.left())),int(max(0,sr.top())),
                                     int(min(self._crop_img_w,sr.right())),int(min(self._crop_img_h,sr.bottom())))
                    self._crop_clbl.setText(f'Crop: ({self._crop_rect[0]},{self._crop_rect[1]})->({self._crop_rect[2]},{self._crop_rect[3]})  {self._crop_rect[2]-self._crop_rect[0]}x{self._crop_rect[3]-self._crop_rect[1]}px')
                return True
        return super().eventFilter(obj,event)

    def _open_raw(self):
        d=QFileDialog.getExistingDirectory(self,'Select Raw Images Folder',str(self._raw_dir))
        if not d: return
        self._raw_dir=Path(d); self._settings.setValue('raw_dir',str(self._raw_dir))
        self._crop_flist.clear()
        # Find all subdirectories containing images (recursive, case-insensitive)
        img_exts = ('.bmp','.jpg','.jpeg','.png','.tif','.tiff')
        for root, dirs, files in os.walk(str(self._raw_dir)):
            if any(f.lower().endswith(img_exts) for f in files):
                rel=Path(root).relative_to(self._raw_dir)
                if str(rel)!='.': self._crop_flist.addItem(str(rel))
        # Check also if root directory itself contains images
        root_files = [f for f in os.listdir(str(self._raw_dir))
                      if os.path.isfile(os.path.join(str(self._raw_dir),f))
                      and f.lower().endswith(img_exts)]
        if root_files:
            self._crop_flist.insertItem(0, '(root)')
        self._crop_lbl.setText(str(self._raw_dir))
        self._status.showMessage(f'Loaded: {self._crop_flist.count()} folders with images')

    def _crop_fchg(self, item):
        if not item: return
        folder_name = item.text()
        if folder_name == '(root)':
            d = self._raw_dir
        else:
            d = self._raw_dir / folder_name
        self._crop_ilist.clear()
        img_exts = ('.bmp','.jpg','.jpeg','.png','.tif','.tiff')
        for f in sorted(d.iterdir()):
            if f.is_file() and f.suffix.lower() in img_exts:
                self._crop_ilist.addItem(f.name)

    def _crop_ichg(self, item):
        if not item or not self._crop_flist.currentItem(): return
        folder_name = self._crop_flist.currentItem().text()
        if folder_name == '(root)':
            d = self._raw_dir
        else:
            d = self._raw_dir / folder_name
        p = d / item.text()
        img = _imread(p)
        if img is None: return
        self._crop_img_w, self._crop_img_h = img.shape[1], img.shape[0]
        self._crop_img_w,self._crop_img_h=img.shape[1],img.shape[0]
        self._crop_view.set_image(img)
        self._crop_rect=None; self._crop_clbl.setText('Drag mouse to select crop region')

    def _set_crop_from_manual(self):
        x1=self._crop_x1.value(); y1=self._crop_y1.value()
        x2=self._crop_x2.value(); y2=self._crop_y2.value()
        if x2<=x1 or y2<=y1:
            QMessageBox.warning(self,'Error','X2>X1 and Y2>Y1 required.'); return
        self._crop_rect=(x1,y1,x2,y2)
        self._crop_clbl.setText(f'Crop: ({x1},{y1})->({x2},{y2})  {x2-x1}x{y2-y1}px')

    def _reuse_last_crop(self):
        crop_file = Path(__file__).parent.parent/'output'/'crops.json'
        if not crop_file.exists():
            QMessageBox.warning(self,'Error','No previous crop saved.'); return
        try:
            with open(crop_file) as f:
                crops = json.load(f)
        except: QMessageBox.warning(self,'Error','Cannot read crop file.'); return
        if not crops:
            QMessageBox.warning(self,'Error','No crops saved yet.'); return
        # Show last 5 saved crops, let user pick
        items = list(crops.items())[-5:]
        msg = 'Select crop region to reuse:\n\n'
        opts = {}
        for i,(k,v) in enumerate(reversed(items)):
            w = v['x2']-v['x1']; h = v['y2']-v['y1']
            label = '{}: ({},{})->({},{}) [{}x{}]'.format(k, v['x1'],v['y1'],v['x2'],v['y2'], w, h)
            msg += '{}: {}\n'.format(chr(65+i), label)
            opts[chr(65+i)] = v
        msg += '\nChoose a letter:'
        reply = QInputDialog.getText(self, 'Reuse Crop', msg)
        if reply[1] and reply[0].strip().upper() in opts:
            c = opts[reply[0].strip().upper()]
            self._crop_rect = (c['x1'],c['y1'],c['x2'],c['y2'])
            self._crop_x1.setValue(c['x1']); self._crop_y1.setValue(c['y1'])
            self._crop_x2.setValue(c['x2']); self._crop_y2.setValue(c['y2'])
            w = c['x2']-c['x1']; h = c['y2']-c['y1']
            self._crop_clbl.setText('Crop: ({},{})->({},{})  {}x{}px'.format(c['x1'],c['y1'],c['x2'],c['y2'],w,h))

    def _choose_crop_out(self):
        d = QFileDialog.getExistingDirectory(self, 'Select Output Directory',
                                              self._crop_out_dir.text())
        if d:
            self._crop_out_dir.setText(d)
            self._settings.setValue('crop_out_dir', d)
            self._settings.setValue('cropped_dir', d)

    def _apply_crop(self):
        if not self._crop_rect or not self._crop_flist.currentItem():
            QMessageBox.warning(self,'Error','Select crop region first.'); return
        x1,y1,x2,y2=self._crop_rect
        fr=self._crop_flist.currentItem().text()
        if fr == '(root)':
            sd = self._raw_dir
        else:
            sd = self._raw_dir / fr

        # Output directory from settings
        od = Path(self._crop_out_dir.text())
        if fr != '(root)':
            od = od / fr
        od.mkdir(parents=True, exist_ok=True)

        crop_key = str(sd.relative_to(self._raw_dir)) if sd != self._raw_dir else '(root)'
        self._save_crop_region(crop_key, x1, y1, x2, y2)

        img_exts = ('.bmp','.jpg','.jpeg','.png','.tif','.tiff')
        bmps = sorted([f for f in sd.iterdir() if f.is_file() and f.suffix.lower() in img_exts])
        if not bmps:
            QMessageBox.warning(self,'Error','No images found in folder.'); return

        # Naming settings
        prefix = self._crop_prefix.text().strip() or 'frame_'
        digits = self._crop_digits.value()
        start_idx = self._crop_start.value()

        pdlg = QProgressDialog(f'Cropping {len(bmps)} images...', 'Cancel', 0, len(bmps), self)
        pdlg.setWindowModality(Qt.WindowModality.WindowModal)
        pdlg.show()
        for i,fp in enumerate(bmps):
            if pdlg.wasCanceled(): break
            img = _imread(fp)
            if img is not None:
                # Build filename: prefix + zero-padded index + original extension
                fname = f'{prefix}{start_idx + i:0{digits}d}.bmp'
                cv2.imwrite(str(od / fname), img[y1:y2, x1:x2])
            pdlg.setValue(i+1)
        pdlg.close()
        self._status.showMessage(f'Cropped {len(bmps)} images -> {od}')
        QMessageBox.information(self,'Done',
            f'{len(bmps)} images cropped ({x2-x1}x{y2-y1}px)\n'
            f'Saved to: {od}\n'
            f'Naming: {prefix}[{start_idx:0{digits}d}-{start_idx+len(bmps)-1:0{digits}d}].bmp')

    def _apply_crop_current(self):
        if not self._crop_rect:
            QMessageBox.warning(self,'Error','Select crop region first.'); return
        if not self._crop_ilist.currentItem() or not self._crop_flist.currentItem():
            QMessageBox.warning(self,'Error','No image selected.'); return
        x1,y1,x2,y2=self._crop_rect
        fr=self._crop_flist.currentItem().text()
        if fr=='(root)':
            sd=self._raw_dir
        else:
            sd=self._raw_dir/fr
        fn=self._crop_ilist.currentItem().text()
        fp=sd/fn
        img=_imread(fp)
        if img is None:
            QMessageBox.warning(self,'Error',f'Cannot read image: {fn}'); return
        od=Path(self._crop_out_dir.text())
        if fr!='(root)':
            od=od/fr
        od.mkdir(parents=True,exist_ok=True)
        crop_key=str(sd.relative_to(self._raw_dir)) if sd!=self._raw_dir else '(root)'
        self._save_crop_region(crop_key,x1,y1,x2,y2)
        out_name=f'cropped_{fn}'
        cv2.imwrite(str(od/out_name),img[y1:y2,x1:x2])
        # Save .txt with crop parameters
        txt_path=od/f'crop_params_{Path(fn).stem}.txt'
        with open(txt_path,'w',encoding='utf-8') as f:
            f.write(f'Crop Parameters\n')
            f.write(f'Source: {fp}\n')
            f.write(f'Crop region: ({x1},{y1}) -> ({x2},{y2})\n')
            f.write(f'Size: {x2-x1}x{y2-y1} px\n')
            f.write(f'Output: {od/out_name}\n')
        self._status.showMessage(f'Cropped: {fn} -> {od/out_name}')
        QMessageBox.information(self,'Done',
            f'Cropped {fn}\n({x2-x1}x{y2-y1}px) -> {od}\n\nParameters saved to:\n{txt_path}')

    def _save_crop_region(self, folder_key, x1, y1, x2, y2):
        crop_file = Path(__file__).parent.parent / 'output' / 'crops.json'
        crop_file.parent.mkdir(parents=True, exist_ok=True)
        crops = {}
        if crop_file.exists():
            try:
                with open(crop_file) as f:
                    crops = json.load(f)
            except: pass
        crops[folder_key] = {'x1': x1, 'y1': y1, 'x2': x2, 'y2': y2}
        with open(crop_file, 'w') as f:
            json.dump(crops, f, indent=2)

    def _apply_crop_all(self):
        if not self._crop_rect or not self._crop_flist.currentItem():
            QMessageBox.warning(self,'Error','Select crop region first.'); return
        x1,y1,x2,y2 = self._crop_rect
        fr = self._crop_flist.currentItem().text()

        # Resolve the actual source folder
        if fr == '(root)':
            src_dir = self._raw_dir
            search_dir = self._raw_dir
        else:
            src_dir = self._raw_dir / fr
            search_dir = src_dir.parent

        img_exts = ('.bmp','.jpg','.jpeg','.png','.tif','.tiff')

        # Collect all sibling folders that have images
        subfolders = []
        for d in sorted(search_dir.iterdir()):
            if not d.is_dir(): continue
            try:
                contents = os.listdir(str(d))
                if any(f.lower().endswith(img_exts) for f in contents
                       if os.path.isfile(os.path.join(str(d), f))):
                    subfolders.append(d)
            except: pass

        if not subfolders:
            QMessageBox.warning(self, 'Error',
                f'No sibling folders with images found in {search_dir}'); return

        # Count total
        total = 0
        for sub in subfolders:
            total += len([f for f in sub.iterdir() if f.is_file() and f.suffix.lower() in img_exts])

        reply = QMessageBox.question(self, 'Confirm',
            f'Apply crop ({x1},{y1})->({x2},{y2}) [{x2-x1}x{y2-y1}px] to {len(subfolders)} folders ({total} images)?\n\n'
            f'Folders: ' + ', '.join([d.name for d in subfolders[:5]]) +
            ('...' if len(subfolders)>5 else ''),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply != QMessageBox.StandardButton.Yes: return

        pdlg = QProgressDialog('Cropping all sibling folders...', 'Cancel', 0, total, self)
        pdlg.setWindowModality(Qt.WindowModality.WindowModal)
        pdlg.show()

        base_out = Path(self._crop_out_dir.text())
        done = 0
        for sub in subfolders:
            try:
                rel = str(sub.relative_to(self._raw_dir))
            except ValueError:
                rel = sub.name
            od = base_out / rel
            od.mkdir(parents=True, exist_ok=True)
            self._save_crop_region(rel.replace(os.sep, '/'), x1, y1, x2, y2)

            for fp in sorted(sub.iterdir()):
                if not fp.is_file() or fp.suffix.lower() not in img_exts:
                    continue
                if pdlg.wasCanceled(): break
                img = _imread(fp)
                if img is not None:
                    cv2.imwrite(str(od / fp.name), img[y1:y2, x1:x2])
                done += 1
                pdlg.setValue(done)
            if pdlg.wasCanceled(): break

        pdlg.close()
        self._settings.setValue('cropped_dir', str(base_out))
        self._status.showMessage(f'Cropped {done} images across {len(subfolders)} folders')
        QMessageBox.information(self,'Done', f'Applied to {len(subfolders)} subfolders\n{done} images cropped ({x2-x1}x{y2-y1}px)')

    # ── Tab 2: Preprocess ──────────────────────────────────

    def _tab_preprocess(self):
        w=QWidget(); l=QVBoxLayout(w)
        top=QHBoxLayout(); top.addWidget(QLabel('Cropped Images Folder:'))
        self._pre_src_lbl=QLabel('(none)'); top.addWidget(self._pre_src_lbl,1)
        top.addWidget(QPushButton('Browse...',clicked=self._open_preprocess))
        l.addLayout(top)

        sp=QSplitter(Qt.Orientation.Horizontal)
        lt=QWidget(); ll=QVBoxLayout(lt); lt.setMaximumWidth(280)
        ll.addWidget(QLabel('Images:')); self._pre_ilist=QListWidget()
        self._pre_ilist.currentRowChanged.connect(self._pre_ichg); ll.addWidget(self._pre_ilist)

        ctrl=QGroupBox('Background Correction'); cl=QVBoxLayout(ctrl)
        self._pre_method=QComboBox(); self._pre_method.addItems(['Gaussian Blur','Temporal Median (all frames)'])
        self._pre_method.currentIndexChanged.connect(self._pre_apply)
        cl.addWidget(self._pre_method)

        self._pre_gauss_lbl=QLabel('Gaussian sigma:'); cl.addWidget(self._pre_gauss_lbl)
        self._pre_gauss=QSpinBox(); self._pre_gauss.setRange(20,200); self._pre_gauss.setValue(80)
        self._pre_gauss.valueChanged.connect(self._pre_apply); cl.addWidget(self._pre_gauss)

        self._pre_temporal_lbl=QLabel('Temporal: using all frames for median'); cl.addWidget(self._pre_temporal_lbl)
        self._pre_temporal_lbl.hide()

        cl.addWidget(QLabel('Output:'))
        self._pre_save_cb=QCheckBox('Save preprocessed images'); self._pre_save_cb.setChecked(True)
        cl.addWidget(self._pre_save_cb)

        sr=QHBoxLayout(); self._pre_out_dir=QLineEdit(
            str(Path(__file__).parent.parent/'output'/'preprocessed'))
        self._pre_out_dir.setReadOnly(True); sr.addWidget(self._pre_out_dir,1)
        pb=QPushButton('...'); pb.setFixedWidth(30); pb.clicked.connect(lambda:
            self._choose_pre_out()); sr.addWidget(pb)
        cl.addLayout(sr)

        cl.addWidget(QLabel('Prefix:')); self._pre_prefix=QLineEdit('pre_')
        cl.addWidget(self._pre_prefix)

        save_btn=QPushButton('Save Preprocessed Images')
        save_btn.setStyleSheet('background:#106040;color:#fff;font-weight:bold;padding:8px')
        save_btn.clicked.connect(self._save_preprocessed); cl.addWidget(save_btn)

        cur_btn=QPushButton('Preprocess Current Image Only')
        cur_btn.setStyleSheet('background:#806020;color:#fff;font-weight:bold;padding:8px')
        cur_btn.clicked.connect(self._preprocess_current); cl.addWidget(cur_btn)

        skip_btn=QPushButton('Skip Save → Go to Detection')
        skip_btn.setStyleSheet('background:#602080;color:#fff;font-weight:bold;padding:8px')
        skip_btn.clicked.connect(self._skip_to_detect); cl.addWidget(skip_btn)
        ll.addWidget(ctrl); ll.addStretch()

        rt=QWidget(); rl=QVBoxLayout(rt)
        self._pre_view=ImageViewer(); rl.addWidget(self._pre_view)
        self._pre_info=QLabel('No image'); rl.addWidget(self._pre_info)

        sp.addWidget(lt); sp.addWidget(rt); sp.setSizes([280,1000])
        l.addWidget(sp)

        self._pre_dir=None; self._pre_img_list=[]; self._pre_idx=0
        self._pre_raw=None; self._pre_corrected=None; self._pre_median_bg=None
        self._pre_skip_btn=skip_btn
        return w

    def _open_preprocess(self):
        d=QFileDialog.getExistingDirectory(self,'Select Cropped Images Folder',
            str(self._settings.value('cropped_dir',str(Path.home()))))
        if not d: return
        self._pre_dir=Path(d); self._pre_src_lbl.setText(str(self._pre_dir))
        self._pre_ilist.clear(); self._pre_img_list=[]
        img_exts=('.bmp','.jpg','.jpeg','.png','.tif','.tiff')
        for f in sorted(self._pre_dir.iterdir()):
            if f.is_file() and f.suffix.lower() in img_exts:
                self._pre_img_list.append(f)
                self._pre_ilist.addItem(f.name)
        self._pre_idx=0; self._pre_median_bg=None
        self._status.showMessage(f'Loaded {len(self._pre_img_list)} images for preprocessing')
        self._pre_ichg(0)

    def _pre_ichg(self, idx):
        if idx<0 or not self._pre_img_list: return
        self._pre_idx=idx
        fp=self._pre_img_list[idx]
        self._pre_raw=cv2.cvtColor(_imread(fp),cv2.COLOR_BGR2GRAY)
        if self._pre_raw is None: return
        self._pre_apply()

    def _pre_apply(self):
        if self._pre_raw is None: return
        gray=self._pre_raw
        method=self._pre_method.currentText()
        try:
            if 'Temporal' in method and self._pre_img_list:
                if self._pre_median_bg is None:
                    QApplication.processEvents()
                    pdlg=QProgressDialog('Building temporal median...','Cancel',0,len(self._pre_img_list),self)
                    pdlg.setWindowModality(Qt.WindowModality.WindowModal); pdlg.show()
                    stack=[]
                    for ii,f in enumerate(self._pre_img_list):
                        if pdlg.wasCanceled(): break
                        img=_imread(f)
                        if img is not None:
                            stack.append(cv2.cvtColor(img,cv2.COLOR_BGR2GRAY) if len(img.shape)==3 else img)
                        pdlg.setValue(ii+1)
                    pdlg.close()
                    if len(stack)>3:
                        self._pre_median_bg=np.median(np.array(stack),axis=0).astype(np.uint8)
                        self._pre_temporal_lbl.setText(f'Temporal median built from {len(stack)} frames')
                        self._pre_temporal_lbl.show()
                if self._pre_median_bg is not None:
                    residual=cv2.subtract(self._pre_median_bg,gray)
                    # Show what the detector sees: residual / median_bg (scaled 0-255)
                    norm = residual.astype(float) / np.clip(self._pre_median_bg, 1, 255)
                    self._pre_corrected = (np.clip(norm / norm.max(), 0, 1) * 255).astype(np.uint8)
                else:
                    bg=cv2.GaussianBlur(gray,(251,251),self._pre_gauss.value())
                    self._pre_corrected=cv2.subtract(bg,gray)
            else:
                bg=cv2.GaussianBlur(gray,(251,251),self._pre_gauss.value())
                self._pre_corrected=cv2.subtract(bg,gray)
        except Exception as e:
            QMessageBox.warning(self,'Error',f'Preprocessing failed: {e}')
            return
        self._pre_view.set_image(self._pre_corrected)
        fp=self._pre_img_list[self._pre_idx]
        self._pre_info.setText(f'{fp.name} | Corrected: {self._pre_corrected.min()}-{self._pre_corrected.max()}')
        # Histogram removed from preprocess tab — moved to Detect & Review tab

    def _save_preprocessed(self):
        if not self._pre_img_list:
            QMessageBox.warning(self,'Error','No images loaded.'); return
        od=Path(self._pre_out_dir.text()); od.mkdir(parents=True,exist_ok=True)
        prefix=self._pre_prefix.text().strip() or 'pre_'
        total=len(self._pre_img_list)
        pdlg=QProgressDialog(f'Saving {total} images...','Cancel',0,total,self)
        pdlg.setWindowModality(Qt.WindowModality.WindowModal); pdlg.show()
        try:
            method=self._pre_method.currentText()
            if 'Temporal' in method and self._pre_median_bg is None:
                pdlg.close(); self._pre_apply(); pdlg.show()
            for i,fp in enumerate(self._pre_img_list):
                if pdlg.wasCanceled(): break
                gray=cv2.cvtColor(_imread(fp),cv2.COLOR_BGR2GRAY)
                if 'Temporal' in method and self._pre_median_bg is not None:
                    residual=cv2.subtract(self._pre_median_bg,gray)
                    norm = residual.astype(float) / np.clip(self._pre_median_bg, 1, 255)
                    out_img = (np.clip(norm / norm.max(), 0, 1) * 255).astype(np.uint8)
                else:
                    bg=cv2.GaussianBlur(gray,(251,251),self._pre_gauss.value())
                    out_img=cv2.subtract(bg,gray)
                cv2.imwrite(str(od/f'{prefix}{fp.stem}.jpg'),out_img,[cv2.IMWRITE_JPEG_QUALITY,90])
                pdlg.setValue(i+1)
        except Exception as e:
            QMessageBox.critical(self,'Error',f'Save failed: {e}')
        pdlg.close()
        self._status.showMessage(f'Saved {total} preprocessed images -> {od}')
        QMessageBox.information(self,'Done',f'{total} images saved to {od}')

    def _preprocess_current(self):
        if not self._pre_img_list:
            QMessageBox.warning(self,'Error','No images loaded.'); return
        idx=self._pre_idx
        if idx<0 or idx>=len(self._pre_img_list):
            QMessageBox.warning(self,'Error','No image selected.'); return
        fp=self._pre_img_list[idx]
        gray=cv2.cvtColor(_imread(fp),cv2.COLOR_BGR2GRAY)
        if gray is None:
            QMessageBox.warning(self,'Error',f'Cannot read image: {fp.name}'); return
        od=Path(self._pre_out_dir.text()); od.mkdir(parents=True,exist_ok=True)
        prefix=self._pre_prefix.text().strip() or 'pre_'
        method=self._pre_method.currentText()
        if 'Temporal' in method:
            if self._pre_median_bg is None:
                self._pre_apply()
            if self._pre_median_bg is not None:
                residual=cv2.subtract(self._pre_median_bg,gray)
                norm=residual.astype(float)/np.clip(self._pre_median_bg,1,255)
                out_img=(np.clip(norm/norm.max(),0,1)*255).astype(np.uint8)
            else:
                bg=cv2.GaussianBlur(gray,(251,251),self._pre_gauss.value())
                out_img=cv2.subtract(bg,gray)
        else:
            bg=cv2.GaussianBlur(gray,(251,251),self._pre_gauss.value())
            out_img=cv2.subtract(bg,gray)
        out_name=f'{prefix}{fp.stem}.jpg'
        cv2.imwrite(str(od/out_name),out_img,[cv2.IMWRITE_JPEG_QUALITY,90])
        # Save .txt with processing parameters
        txt_path=od/f'{prefix}{fp.stem}_params.txt'
        with open(txt_path,'w',encoding='utf-8') as f:
            f.write(f'Preprocess Parameters\n')
            f.write(f'Source: {fp}\n')
            f.write(f'Method: {method}\n')
            if 'Temporal' in method:
                f.write(f'Median background: {len(self._pre_img_list)} frames\n')
            else:
                f.write(f'Gaussian sigma: {self._pre_gauss.value()}\n')
            f.write(f'Kernel size: 251x251\n')
            f.write(f'Output: {od/out_name}\n')
        self._status.showMessage(f'Preprocessed: {fp.name} -> {od/out_name}')
        QMessageBox.information(self,'Done',
            f'Preprocessed {fp.name}\nMethod: {method}\n-> {od}\n\nParameters saved to:\n{txt_path}')

    def _skip_to_detect(self):
        if not self._pre_img_list:
            QMessageBox.warning(self,'Error','No images loaded.'); return
        od=Path(self._pre_out_dir.text()); od.mkdir(parents=True,exist_ok=True)
        prefix=self._pre_prefix.text().strip() or 'pre_'
        total=len(self._pre_img_list)
        pdlg=QProgressDialog(f'Processing {total} images...','Cancel',0,total,self)
        pdlg.setWindowModality(Qt.WindowModality.WindowModal); pdlg.show()
        try:
            method=self._pre_method.currentText()
            if 'Temporal' in method and self._pre_median_bg is None:
                stack=[]
                for fp2 in self._pre_img_list:
                    img=_imread(fp2)
                    if img is not None:
                        stack.append(cv2.cvtColor(img,cv2.COLOR_BGR2GRAY) if len(img.shape)==3 else img)
                if len(stack)>3:
                    self._pre_median_bg=np.median(np.array(stack),axis=0).astype(np.uint8)
            for i,fp in enumerate(self._pre_img_list):
                if pdlg.wasCanceled(): break
                gray=cv2.cvtColor(_imread(fp),cv2.COLOR_BGR2GRAY)
                if 'Temporal' in method and self._pre_median_bg is not None:
                    residual=cv2.subtract(self._pre_median_bg,gray)
                    norm = residual.astype(float) / np.clip(self._pre_median_bg, 1, 255)
                    out_img = (np.clip(norm / norm.max(), 0, 1) * 255).astype(np.uint8)
                else:
                    bg=cv2.GaussianBlur(gray,(251,251),self._pre_gauss.value())
                    out_img=cv2.subtract(bg,gray)
                cv2.imwrite(str(od/f'{prefix}{fp.stem}.jpg'),out_img,[cv2.IMWRITE_JPEG_QUALITY,90])
                pdlg.setValue(i+1); QApplication.processEvents()
        except Exception as e:
            QMessageBox.critical(self,'Error',f'Skip-to-detect failed: {e}')
        pdlg.close()
        self._cropped_dir=od; self._det_lbl.setText(str(od))
        self._det_ilist.clear(); self._img_list=[]
        for f in sorted(od.iterdir()):
            if f.is_file() and f.suffix.lower() in ('.jpg','.jpeg','.png','.bmp','.tif'):
                self._img_list.append(f); self._det_ilist.addItem(f.name)
        self._det_dir=Path(__file__).parent.parent/'output'/'temporary'/'detections'
        self._det_dir.mkdir(parents=True,exist_ok=True)
        self._detections={}
        for jf in self._det_dir.glob('*.json'):
            fname=jf.stem+'.bmp'
            try:
                with open(jf) as f: self._detections[fname]=json.load(f).get('droplets',[])
            except: pass
        self._status.showMessage(f'Ready for detection ({len(self._img_list)} images)')
        if len(self._img_list)>0:
            self._tabs.setCurrentIndex(2)

    # ── Tab 3: Detect & Review ──────────────────────────────

    def _tab_detect(self):
        w=QWidget(); l=QHBoxLayout(w)
        lt=QWidget(); ll=QVBoxLayout(lt); lt.setMaximumWidth(280)
        ll.addWidget(QLabel('Cropped images:'))
        fr=QHBoxLayout(); self._det_lbl=QLabel('(none)'); fr.addWidget(self._det_lbl,1)
        fr.addWidget(QPushButton('...',clicked=self._open_cropped)); ll.addLayout(fr)
        self._det_ilist=QListWidget()
        self._det_ilist.currentRowChanged.connect(self._det_ichg); ll.addWidget(self._det_ilist)
        ll.addWidget(QLabel('Parameters:'))
        self._det_par={}
        # Default params calibrated against 490 expert annotations
        # D_eq ratio=1.01x, D32 ratio=1.01x (all quantiles within 0.97-1.02x)
        for nm,dv,rg in [('min_distance',5,(2,20)),('min_area',10,(3,200)),('max_area',350,(100,3000)),
                         ('max_aspect',2.0,(1.5,5.0)),('peak_threshold',0.04,(0.01,0.5)),
                         ('seed_threshold',0.05,(0.01,0.2)),('bg_sigma',80,(20,200))]:
            rw=QHBoxLayout(); rw.addWidget(QLabel(nm))
            sp=QDoubleSpinBox() if isinstance(dv,float) else QSpinBox()
            sp.setRange(int(rg[0]) if isinstance(dv,int) else rg[0],int(rg[1]) if isinstance(dv,int) else rg[1])
            if isinstance(dv,float): sp.setDecimals(3); sp.setSingleStep(0.01)
            sp.setValue(dv); rw.addWidget(sp); ll.addLayout(rw); self._det_par[nm]=sp
        ll.addWidget(QLabel('Scale:')); self._det_sc=QDoubleSpinBox()
        self._det_sc.setRange(1,10000); self._det_sc.setValue(UM_PER_PX_DEFAULT)
        self._det_sc.setSuffix(' um/px'); ll.addWidget(self._det_sc)
        rb_one=QPushButton('Detect Current Image Only')
        rb_one.setStyleSheet('background:#4060a0;color:#fff;font-weight:bold;padding:8px')
        rb_one.setToolTip('Detect droplets on the currently displayed image using Temporal Median (if checkbox checked)')
        rb_one.clicked.connect(self._run_det_one); ll.addWidget(rb_one)
        rb=QPushButton('Run Detection on All Images')
        rb.setStyleSheet('background:#106040;color:#fff;font-weight:bold;padding:8px')
        rb.clicked.connect(self._run_det); ll.addWidget(rb)
        self._det_pb=QProgressBar(); ll.addWidget(self._det_pb)

        # Temporal background subtraction checkbox
        self._det_temporal = QCheckBox('Temporal Background Subtraction (median of frames)')
        self._det_temporal.setToolTip(
            'Use temporal median across frames to remove static background.\n'
            'Best for PIV image sequences where droplets move between frames.\n'
            'Disable for single images or images from different positions.')
        self._det_temporal.setChecked(True)
        ll.addWidget(self._det_temporal)

        # Settings: CPU cores
        st_box = QGroupBox('Settings')
        st_lay = QVBoxLayout(st_box)
        cr_row = QHBoxLayout()
        cr_row.addWidget(QLabel('CPU Cores:'))
        total_cores = multiprocessing.cpu_count()
        self._det_cores = QSpinBox()
        self._det_cores.setRange(1, total_cores)
        saved_cores = int(self._settings.value('det_cores', str(total_cores)))
        self._det_cores.setValue(min(saved_cores, total_cores))
        self._det_cores.setToolTip(f'Number of CPU cores for batch detection (1-{total_cores}, default: {total_cores})')
        cr_row.addWidget(self._det_cores)
        cr_row.addStretch()
        st_lay.addLayout(cr_row)
        ll.addWidget(st_box)

        ll.addWidget(QLabel('Annotation Mode:')); self._det_mlbl=QLabel('🖱 View — drag=pan, click=select, scroll=zoom')
        ll.addWidget(self._det_mlbl)
        mg=QButtonGroup(self); mg.setExclusive(True)
        mr=QHBoxLayout()
        vbtn=QPushButton('🖱 View'); vbtn.setCheckable(True); vbtn.setChecked(True)
        vbtn.toggled.connect(lambda v: v and self._det_view.set_mode('view'))
        vbtn.toggled.connect(lambda v: v and self._det_mlbl.setText('🖱 View — drag=pan, click=select, scroll=zoom'))
        mr.addWidget(vbtn); mg.addButton(vbtn)
        abtn=QPushButton('+ Add'); abtn.setCheckable(True)
        abtn.toggled.connect(lambda v: v and self._det_view.set_mode('add'))
        abtn.toggled.connect(lambda v: v and self._det_mlbl.setText(
            '+ Add Mode — hold left button → trace droplet edge → release'))
        mr.addWidget(abtn); mg.addButton(abtn)
        dbtn=QPushButton('✕ Delete'); dbtn.setCheckable(True)
        dbtn.toggled.connect(lambda v: v and self._det_view.set_mode('delete'))
        dbtn.toggled.connect(lambda v: v and self._det_mlbl.setText(
            '✕ Delete Mode — click on a droplet to remove it'))
        mr.addWidget(dbtn); mg.addButton(dbtn)
        ll.addLayout(mr)
        sbtn=QPushButton('💾 Save'); sbtn.clicked.connect(self._save_det); ll.addWidget(sbtn)
        ll.addStretch()
        rt=QWidget(); rl=QVBoxLayout(rt)
        self._det_view=ImageViewer(); rl.addWidget(self._det_view)
        self._det_info=QLabel('No image'); rl.addWidget(self._det_info)
        self._det_hist=QLabel()
        self._det_hist.setToolTip('Particle size distribution for current image')
        self._det_hist.setMinimumHeight(200)
        self._det_hist.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._det_hist.setStyleSheet('background:#f8f8f8;border:2px solid #4472C4;margin-top:6px;padding:4px')
        self._det_hist.setText('')
        rl.addWidget(self._det_hist)
        l.addWidget(lt); l.addWidget(rt,1)
        self._det_dir=None
        return w

    def _open_cropped(self):
        d=QFileDialog.getExistingDirectory(self,'Select Cropped Images Folder',
            str(self._settings.value('cropped_dir',str(Path.home()))))
        if not d: return
        self._cropped_dir=Path(d); self._det_lbl.setText(str(self._cropped_dir))
        self._det_ilist.clear(); self._img_list=sorted([f for f in self._cropped_dir.iterdir() if f.suffix.lower()=='.bmp'])
        for f in self._img_list: self._det_ilist.addItem(f.name)
        self._settings.setValue('cropped_dir',str(d))
        # Check for existing detection files
        self._det_dir=Path(__file__).parent.parent/'output'/'temporary'/'detections'
        self._det_dir.mkdir(parents=True,exist_ok=True)
        self._det_dir.mkdir(parents=True,exist_ok=True)
        self._detections={}
        for jf in self._det_dir.glob('*.json'):
            fname=jf.stem+'.bmp'
            try:
                with open(jf) as f: self._detections[fname]=json.load(f).get('droplets',[])
            except: pass

    def _det_ichg(self, idx):
        try:
            if idx<0 or not self._img_list: return
            fp=self._img_list[idx]
            img=_imread(fp)
            if img is None: return
            self._img_idx=idx; drops=self._detections.get(fp.name,[])
            self._det_view.set_image(img); self._det_view.set_droplets(drops)
            self._det_info.setText(f'#{idx+1}/{len(self._img_list)}: {fp.name}  |  {len(drops)} droplets  |  {img.shape[1]}x{img.shape[0]}')
            self._det_show_histogram(drops)
        except Exception as e:
            pass

    def _run_det(self):
        if not self._img_list: QMessageBox.warning(self,'Error','No images.'); return
        if getattr(self,'_w',None) and self._w.isRunning():
            self._w.quit(); self._w.wait()
        params={k:v.value() for k,v in self._det_par.items()}
        cores = self._det_cores.value()
        self._settings.setValue('det_cores', str(cores))
        self._det_pb.setMaximum(len(self._img_list)); self._det_pb.setValue(0)
        use_temp = self._det_temporal.isChecked()
        self._det_det_worker = DetectWorker(self._img_list, params, num_workers=cores, use_temporal=use_temp)
        self._w = self._det_det_worker
        self._w.progress.connect(lambda i,t: self._det_pb.setValue(i))
        self._w.status.connect(lambda s: self._status.showMessage(s))
        self._w.error.connect(lambda s: QMessageBox.warning(self,'Error',s))
        self._w.finished.connect(self._on_det_done); self._w.start()

    def _run_det_one(self):
        try:
            if not self._img_list or self._img_idx < 0 or self._img_idx >= len(self._img_list):
                QMessageBox.warning(self,'Error','No image selected.'); return
            fp = self._img_list[self._img_idx]
            params={k:v.value() for k,v in self._det_par.items()}

            self._status.showMessage(f'Detecting: {fp.name}...')
            QApplication.processEvents()

            img = _imread(fp)
            if img is None:
                QMessageBox.warning(self,'Error',f'Cannot read image: {fp.name}'); return
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

            median_bg = None
            if self._det_temporal.isChecked() and len(self._img_list) >= 3:
                cached = getattr(self, '_cached_median_bg', None)
                if cached is not None:
                    median_bg = cached
                else:
                    self._status.showMessage('Building temporal median background...')
                    QApplication.processEvents()
                    median_bg = _build_median_bg(self._img_list)
                    if median_bg is not None:
                        self._cached_median_bg = median_bg

            if median_bg is not None:
                result = detect_droplets(gray, median_bg=median_bg, **params)
                self._det_view.set_median_bg(median_bg)
            else:
                result = detect_droplets(gray, **params)

            self._detections[fp.name] = result
            self._status.showMessage(f'Done: {len(result)} droplets')
            if len(result) == 0:
                self._status.showMessage(
                    'No droplets detected. Check: 1) Image loaded correctly? '
                    '2) Temporal checkbox matches your data? '
                    '3) Parameters appropriate for your images?')
            self._det_ichg(self._img_idx)
            # Show histogram after detection
            self._det_show_histogram(result)
        except Exception as e:
            tb = traceback.format_exc()
            with open('output/crash_log.txt', 'w', encoding='utf-8') as f:
                f.write(f'=== CRASH {datetime.datetime.now()} ===\n{tb}\n')
            QMessageBox.critical(self, 'Detection Crashed',
                f'{e}\n\nFull traceback written to: output/crash_log.txt')

    def _on_det_done(self, results):
        try:
            self._detections=results
            total=sum(len(v) for v in results.values())
            self._status.showMessage(f'Done: {total:,} droplets in {len(results)} images')
            # Cache median background for single-image detection + viewer
            mb = getattr(self._det_det_worker, 'median_bg', None)
            if mb is not None:
                self._cached_median_bg = mb
                self._det_view.set_median_bg(mb)
            self._det_ichg(self._img_idx)
            # Update histogram for current image
            fn = self._img_list[self._img_idx].name if self._img_list else None
            if fn in results:
                self._det_show_histogram(results[fn])
        except Exception as e:
            pass

    def _det_show_histogram(self, drops):
        try:
            if not drops:
                self._det_hist.setText('No droplets detected')
                return
            dias=[d['axes'][0] for d in drops]
            _lazy_import()
            fig=Figure(figsize=(5,2)); ax=fig.subplots()
            ax.hist(dias,bins=min(60,max(10,len(dias)//3)),color='steelblue',edgecolor='white',alpha=0.85)
            ax.axvline(np.mean(dias),color='red',ls='--',lw=1,label=f'Mean={np.mean(dias):.1f}px')
            ax.set_xlabel('D_eq (px)'); ax.set_ylabel('Count')
            ax.set_title(f'Size Distribution (n={len(drops)})',fontsize=10)
            ax.legend(fontsize=7)
            fig.tight_layout()
            od=Path(__file__).parent.parent/'output'
            od.mkdir(parents=True,exist_ok=True)
            fig.savefig(str(od/'_det_hist.png'),dpi=100)
            pm=QPixmap(str(od/'_det_hist.png'))
            if not pm.isNull():
                scaled=pm.scaled(self._det_hist.size(),
                    Qt.AspectRatioMode.KeepAspectRatio,
                    Qt.TransformationMode.SmoothTransformation)
                self._det_hist.setPixmap(scaled)
        except Exception:
            # Histogram is best-effort — main detection results are preserved
            pass

    def _save_det(self):
        if not self._img_list: return
        fn=self._img_list[self._img_idx].name
        drops=self._det_view._droplets
        self._detections[fn]=drops
        if self._det_dir is None:
            self._det_dir=Path(__file__).parent.parent/'output'/'temporary'/'detections'
        self._det_dir.mkdir(parents=True,exist_ok=True)
        with open(self._det_dir/(Path(fn).stem+'.json'),'w') as f:
            json.dump({'droplets':drops,'count':len(drops)},f,indent=2)
        self._det_view._modified=False
        self._status.showMessage(f'Saved: {fn} ({len(drops)} droplets)')

    # ── Tab 3: Export ───────────────────────────────────────

    def _tab_export(self):
        w=QWidget(); l=QVBoxLayout(w)
        r0=QHBoxLayout(); r0.addWidget(QLabel('Export to:'))
        self._exp_out_dir=QLineEdit(str(Path(__file__).parent.parent/'output'/'reports')); self._exp_out_dir.setReadOnly(True)
        self._exp_out_dir.setStyleSheet('color:#2060a0;font-size:11px')
        r0.addWidget(self._exp_out_dir,1)
        eb=QPushButton('Browse...'); eb.setFixedWidth(65); eb.clicked.connect(self._choose_exp_out); r0.addWidget(eb)
        l.addLayout(r0)
        r1=QHBoxLayout(); r1.addWidget(QLabel('Scale:'))
        self._exp_sc=QDoubleSpinBox(); self._exp_sc.setRange(1,10000)
        self._exp_sc.setValue(UM_PER_PX_DEFAULT); self._exp_sc.setSuffix(' um/px')
        r1.addWidget(self._exp_sc); r1.addStretch(); l.addLayout(r1)
        r2=QHBoxLayout()
        for txt,slot,sty in [('Export Annotated Images',self._exp_ann,'#106040'),
            ('Export Excel + Plots',self._exp_stats,'#204080'),
            ('Export All',self._exp_all,'#602080;font-weight:bold')]:
            b=QPushButton(txt); b.clicked.connect(slot)
            b.setStyleSheet(f'background:{sty};color:#fff;padding:10px;font-size:13px'); r2.addWidget(b)
        l.addLayout(r2)
        self._exp_pb=QProgressBar(); l.addWidget(self._exp_pb)
        self._exp_preview=QLabel(); self._exp_preview.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._exp_preview.setMinimumHeight(400)
        self._exp_preview.setStyleSheet('background:#fff;border:1px solid #ccc')
        l.addWidget(self._exp_preview,1)
        return w

    def _choose_exp_out(self):
        d = QFileDialog.getExistingDirectory(self, 'Select Export Directory', self._exp_out_dir.text())
        if d:
            self._exp_out_dir.setText(d)

    def _check_export(self):
        if not self._cropped_dir or not self._img_list:
            QMessageBox.warning(self,'Error','Open cropped images in Detect tab first.'); return None,None,None
        det_dir=self._det_dir
        if not det_dir or not list(det_dir.glob('*.json')):
            QMessageBox.warning(self,'Error','Run detection and save first.'); return None,None,None
        out=Path(self._settings.value('export_dir',str(self._cropped_dir.parent/'results')))
        out.mkdir(parents=True,exist_ok=True)
        self._settings.setValue('export_dir',str(out))
        return self._cropped_dir,det_dir,out

    def _exp_ann(self):
        r=self._check_export()
        src,det,_=r; out=Path(self._exp_out_dir.text())
        if not src: return
        ann=out/'annotated'; ann.mkdir(exist_ok=True)
        cols=[(0,255,0),(255,255,0),(0,255,255),(255,0,255),(0,128,255),(255,128,0),(128,255,0),(0,0,255)]
        jfs=sorted(det.glob('*.json')); self._exp_pb.setMaximum(len(jfs))
        for i,jf in enumerate(jfs):
            bmp=src/(jf.stem+'.bmp'); img=_imread(bmp)
            if img is None: continue
            with open(jf) as f:
                dets=json.load(f)
            for d in dets.get('droplets',[]):
                cv2.polylines(img,[np.array(d['contour'],np.int32)],True,cols[d['id']%8],1)
            cv2.imwrite(str(ann/(jf.stem+'.jpg')),img,[cv2.IMWRITE_JPEG_QUALITY,88])
            self._exp_pb.setValue(i+1); QApplication.processEvents()
        self._status.showMessage(f'Annotated images: {ann}')
        QMessageBox.information(self,'Done',f'{len(jfs)} annotated images ->\n{ann}')

    def _exp_stats(self):
        r=self._check_export()
        src,det,_=r; out=Path(self._exp_out_dir.text())
        if not src: return
        um=self._exp_sc.value(); self._exp_pb.setMaximum(100)
        self._ew=ExportWorker(str(src),str(det),str(out),um)
        self._ew.progress.connect(lambda i,t: self._exp_pb.setValue(int(i/t*100)))
        self._ew.status.connect(lambda s: self._status.showMessage(s))
        self._ew.finished.connect(self._on_exp_done); self._ew.start()

    def _exp_all(self): self._exp_ann(); self._exp_stats()

    def _on_exp_done(self, out_dir):
        self._status.showMessage(f'Export: {out_dir}')
        pp=Path(out_dir)/'size_distribution.png'
        if pp.exists():
            self._exp_preview.setPixmap(QPixmap(str(pp)).scaled(
                self._exp_preview.width(),self._exp_preview.height(),
                Qt.AspectRatioMode.KeepAspectRatio,Qt.TransformationMode.SmoothTransformation))
        QMessageBox.information(self,'Done',f'Excel + Plots saved to:\n{out_dir}')

    def closeEvent(self, e):
        # Check for unsaved changes in review tab
        has_unsaved = False
        if hasattr(self, '_det_view') and hasattr(self._det_view, '_modified'):
            has_unsaved = self._det_view._modified
        if has_unsaved:
            reply = QMessageBox.question(self, 'Unsaved Changes',
                'You have unsaved detection edits. Save before exit?',
                QMessageBox.StandardButton.Save |
                QMessageBox.StandardButton.Discard |
                QMessageBox.StandardButton.Cancel)
            if reply == QMessageBox.StandardButton.Save:
                self._save_det()
            elif reply == QMessageBox.StandardButton.Cancel:
                e.ignore(); return
        self._settings.setValue('geometry', self.saveGeometry())
        super().closeEvent(e)

# ── Entry ──────────────────────────────────────────────────────

def main():
    import traceback, datetime
    try:
        QApplication.setHighDpiScaleFactorRoundingPolicy(Qt.HighDpiScaleFactorRoundingPolicy.PassThrough)
        app=QApplication(sys.argv); app.setApplicationName('ImageIdentification')
        _lazy_import()
        w=MainWindow()
        # Center on primary screen
        screen = QApplication.primaryScreen()
        center = screen.availableGeometry().center()
        w.resize(max(1200, min(1400, screen.availableGeometry().width()*8//10)),
                 max(750, min(900, screen.availableGeometry().height()*8//10)))
        w.move(center.x() - w.width()//2, center.y() - w.height()//2)
        w.show()
        sys.exit(app.exec())
    except Exception as e:
        log = os.path.join(os.path.dirname(sys.executable) if getattr(sys,'frozen',False) else '.', 'main_crash.log')
        with open(log, 'w', encoding='utf-8') as f:
            f.write(f'=== main() CRASH {datetime.datetime.now()} ===\n{traceback.format_exc()}\n')

if __name__=='__main__': main()
